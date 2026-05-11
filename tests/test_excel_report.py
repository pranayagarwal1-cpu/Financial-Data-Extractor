"""Tests for Excel output including the Quality Report sheet."""

import json
import tempfile
from pathlib import Path

import openpyxl
import pytest

from utils.excel_writer import save_to_excel


class TestQualityReportSchema:
    def test_quality_report_sheet_present(self):
        data = {
            "statement_type": "income_statement",
            "title": "Test Income Statement",
            "periods": ["2024-12-31"],
            "sections": [
                {"name": "Revenue", "rows": [
                    {"label": "Sales", "values": ["100"], "is_subtotal": False},
                ]}
            ],
        }
        report_metadata = {
            "pdf_name": "test.pdf",
            "statement_type": "income_statement",
            "run_id": "abc123",
            "timestamp": "20260101_120000",
            "overall_passed": True,
            "scores": {
                "coverage": 8.0,
                "format": 9.0,
                "structure": 8.5,
                "content": 9.0,
                "overall": 8.7,
            },
            "findings": [
                {"check_id": "A1", "status": "PASS", "message": "All sections present"},
                {"check_id": "C6", "status": "ADVISORY", "message": "2 subtotals have no children"},
            ],
            "guardrail_flags": [],
            "feedback": "Extraction passed with minor advisories.",
            "retry_count": 1,
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            save_to_excel(data, tmp.name, report_metadata=report_metadata)
            wb = openpyxl.load_workbook(tmp.name)

        assert "Quality Report" in wb.sheetnames
        ws = wb["Quality Report"]

        # Title
        assert ws["A1"].value == "Extraction Quality Report"

        # Metadata
        assert ws["A2"].value == "Statement Type:"
        assert ws["B2"].value == "Income Statement"

        # Overall status
        assert ws["A7"].value == "PASS"

        # Score breakdown
        assert ws["A9"].value == "Section"
        assert ws["A10"].value == "Coverage"
        assert ws["B10"].value == 8.0

        # Overall row
        overall_row = 14
        assert ws["A" + str(overall_row)].value == "Overall"
        assert ws["B" + str(overall_row)].value == 8.7

        # Guardrail flags
        assert any(cell.value == "Guardrail Flags:" for cell in ws["A"])

        # Retry history
        assert any(cell.value == "Retry History:" for cell in ws["A"])

    def test_quality_report_fail_styling(self):
        data = {
            "statement_type": "balance_sheet",
            "title": "Test Balance Sheet",
            "periods": ["2024-12-31"],
            "sections": [],
        }
        report_metadata = {
            "pdf_name": "bad.pdf",
            "statement_type": "balance_sheet",
            "run_id": "run_1",
            "timestamp": "20260101_120000",
            "overall_passed": False,
            "scores": {
                "coverage": 4.0,
                "format": 5.0,
                "structure": 3.0,
                "content": 5.0,
                "overall": 4.1,
            },
            "findings": [
                {"check_id": "A1", "status": "FAIL", "message": "Missing sections"},
            ],
            "guardrail_flags": ["quality_degraded"],
            "feedback": "Extraction failed.",
            "retry_count": 2,
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            save_to_excel(data, tmp.name, report_metadata=report_metadata)
            wb = openpyxl.load_workbook(tmp.name)

        ws = wb["Quality Report"]
        assert ws["A7"].value == "FAIL"

    def test_no_quality_report_when_metadata_missing(self):
        data = {
            "statement_type": "income_statement",
            "title": "Test",
            "periods": ["2024-12-31"],
            "sections": [],
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            save_to_excel(data, tmp.name)
            wb = openpyxl.load_workbook(tmp.name)

        assert "Quality Report" not in wb.sheetnames
