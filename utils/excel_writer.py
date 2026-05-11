import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import utils


# ---------------------------------------------------------------------------
# Phase 4 — Quality Report Sheet
# ---------------------------------------------------------------------------

def _write_quality_report_sheet(wb: openpyxl.Workbook, metadata: dict):
    """
    Append a "Quality Report" sheet to the workbook.

    Layout mirrors the PRODUCTION_PLAN spec:
    - Title, metadata, overall PASS/FAIL
    - 4-section score breakdown
    - Programmatic check statuses
    - Guardrail flags
    - Evaluator feedback
    - Retry history
    - Detailed findings table
    """
    ws = wb.create_sheet(title="Quality Report")

    # Reusable styles
    dark_blue = PatternFill("solid", start_color="1F4E79")
    white_font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    normal_font = Font(name="Arial", size=10)
    small_font = Font(name="Arial", size=9)
    pass_fill = PatternFill("solid", start_color="C6EFCE")
    pass_font = Font(name="Arial", bold=True, size=14, color="006100")
    fail_fill = PatternFill("solid", start_color="FFC7CE")
    fail_font = Font(name="Arial", bold=True, size=14, color="9C0006")
    advisory_fill = PatternFill("solid", start_color="FFEB9C")
    advisory_font = Font(name="Arial", size=10, color="9C5700")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    # Helper to write a row of label + value
    def _meta_row(row_num: int, label: str, value: str, label_fill=None, value_font=None):
        ws.cell(row=row_num, column=1, value=label).font = normal_font
        ws.cell(row=row_num, column=2, value=value).font = value_font or normal_font
        ws.cell(row=row_num, column=1).alignment = left_align
        ws.cell(row=row_num, column=2).alignment = left_align

    # Row 1 — Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    title_cell = ws.cell(row=1, column=1, value="Extraction Quality Report")
    title_cell.font = white_font
    title_cell.fill = dark_blue
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 24

    # Row 2 — Metadata
    pdf_name = metadata.get("pdf_name", "")
    statement_type = metadata.get("statement_type", "")
    run_id = metadata.get("run_id", "")
    timestamp = metadata.get("timestamp", "")
    _meta_row(2, "Statement Type:", statement_type.replace("_", " ").title())
    _meta_row(3, "PDF Name:", pdf_name)
    _meta_row(4, "Run ID:", run_id)
    _meta_row(5, "Timestamp:", timestamp)

    # Row 7 — Overall Status
    overall_passed = metadata.get("overall_passed", False)
    status_text = "PASS" if overall_passed else "FAIL"
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=4)
    status_cell = ws.cell(row=7, column=1, value=status_text)
    status_cell.font = pass_font if overall_passed else fail_font
    status_cell.fill = pass_fill if overall_passed else fail_fill
    status_cell.alignment = center_align
    ws.row_dimensions[7].height = 28

    # Row 9 — Score Breakdown header
    ws.cell(row=9, column=1, value="Section").font = header_font
    ws.cell(row=9, column=1).fill = dark_blue
    ws.cell(row=9, column=2, value="Score").font = header_font
    ws.cell(row=9, column=2).fill = dark_blue
    ws.cell(row=9, column=3, value="Weight").font = header_font
    ws.cell(row=9, column=3).fill = dark_blue
    ws.cell(row=9, column=4, value="Status").font = header_font
    ws.cell(row=9, column=4).fill = dark_blue

    scores = metadata.get("scores", {})
    weights = {"coverage": "20%", "format": "20%", "structure": "30%", "content": "30%"}
    row = 10
    for section in ("coverage", "format", "structure", "content"):
        score = scores.get(section, 0)
        passed_sec = score >= 6.0
        ws.cell(row=row, column=1, value=section.replace("_", " ").title()).font = normal_font
        score_cell = ws.cell(row=row, column=2, value=score)
        score_cell.font = normal_font
        score_cell.alignment = right_align
        ws.cell(row=row, column=3, value=weights.get(section, "")).font = normal_font
        ws.cell(row=row, column=3).alignment = center_align
        status_cell = ws.cell(row=row, column=4, value="PASS" if passed_sec else "FAIL")
        status_cell.font = pass_font if passed_sec else fail_font
        status_cell.fill = pass_fill if passed_sec else fail_fill
        row += 1

    # Overall score row
    ws.cell(row=row, column=1, value="Overall").font = Font(name="Arial", bold=True, size=10)
    overall_score = scores.get("overall", 0)
    ws.cell(row=row, column=2, value=overall_score).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=row, column=2).alignment = right_align
    ws.cell(row=row, column=3, value="100%").font = normal_font
    overall_status_cell = ws.cell(row=row, column=4, value="PASS" if overall_passed else "FAIL")
    overall_status_cell.font = pass_font if overall_passed else fail_font
    overall_status_cell.fill = pass_fill if overall_passed else fail_fill
    row += 2

    # Programmatic Checks
    ws.cell(row=row, column=1, value="Check").font = header_font
    ws.cell(row=row, column=1).fill = dark_blue
    ws.cell(row=row, column=2, value="Status").font = header_font
    ws.cell(row=row, column=2).fill = dark_blue
    ws.cell(row=row, column=3, value="Detail").font = header_font
    ws.cell(row=row, column=3).fill = dark_blue
    header_row = row
    row += 1

    findings = metadata.get("findings", [])
    # Aggregate by check_id — show worst status
    check_status: dict[str, str] = {}
    check_detail: dict[str, str] = {}
    for f in findings:
        cid = f.get("check_id", "")
        status = f.get("status", "PASS")
        msg = f.get("message", "")
        if cid not in check_status or status == "FAIL" or (status == "ADVISORY" and check_status[cid] == "PASS"):
            check_status[cid] = status
            check_detail[cid] = msg

    for cid, status in sorted(check_status.items()):
        ws.cell(row=row, column=1, value=cid).font = normal_font
        st_cell = ws.cell(row=row, column=2, value=status)
        if status == "PASS":
            st_cell.font = Font(name="Arial", size=10, color="006100")
        elif status == "FAIL":
            st_cell.font = Font(name="Arial", size=10, color="9C0006")
        else:
            st_cell.font = advisory_font
            st_cell.fill = advisory_fill
        ws.cell(row=row, column=3, value=check_detail.get(cid, "")).font = small_font
        ws.cell(row=row, column=3).alignment = left_align
        row += 1

    if not check_status:
        ws.cell(row=row, column=1, value="All checks passed").font = Font(name="Arial", size=10, color="006100")
        row += 1

    row += 1

    # Guardrail Flags
    ws.cell(row=row, column=1, value="Guardrail Flags:").font = Font(name="Arial", bold=True, size=10)
    row += 1
    guardrail_flags = metadata.get("guardrail_flags", [])
    if guardrail_flags:
        for flag in guardrail_flags:
            cell = ws.cell(row=row, column=1, value=flag)
            cell.font = Font(name="Arial", size=10, color="9C0006")
            row += 1
    else:
        ws.cell(row=row, column=1, value="None").font = Font(name="Arial", size=10, color="006100")
        row += 1

    row += 1

    # Evaluator Feedback
    ws.cell(row=row, column=1, value="Evaluator Feedback:").font = Font(name="Arial", bold=True, size=10)
    row += 1
    feedback = metadata.get("feedback", "")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    fb_cell = ws.cell(row=row, column=1, value=feedback)
    fb_cell.font = small_font
    fb_cell.alignment = left_align
    ws.row_dimensions[row].height = 60
    row += 2

    # Retry History
    retry_count = metadata.get("retry_count", 0)
    ws.cell(row=row, column=1, value="Retry History:").font = Font(name="Arial", bold=True, size=10)
    row += 1
    ws.cell(row=row, column=1, value=f"Attempts used: {retry_count}").font = normal_font
    row += 2

    # Detailed Findings (only failures / advisories)
    failed_findings = [f for f in findings if f.get("status") in ("FAIL", "ADVISORY")]
    if failed_findings:
        ws.cell(row=row, column=1, value="Detailed Findings:").font = Font(name="Arial", bold=True, size=10)
        row += 1
        ws.cell(row=row, column=1, value="Check").font = header_font
        ws.cell(row=row, column=1).fill = dark_blue
        ws.cell(row=row, column=2, value="Status").font = header_font
        ws.cell(row=row, column=2).fill = dark_blue
        ws.cell(row=row, column=3, value="Message").font = header_font
        ws.cell(row=row, column=3).fill = dark_blue
        row += 1
        for f in failed_findings:
            ws.cell(row=row, column=1, value=f.get("check_id", "")).font = normal_font
            st_cell = ws.cell(row=row, column=2, value=f.get("status", ""))
            if f.get("status") == "FAIL":
                st_cell.font = Font(name="Arial", size=10, color="9C0006")
            else:
                st_cell.font = advisory_font
                st_cell.fill = advisory_fill
            ws.cell(row=row, column=3, value=f.get("message", "")).font = small_font
            ws.cell(row=row, column=3).alignment = left_align
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            row += 1

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 14


def save_to_excel(data: dict, output_path: str, include_coa_columns: bool = True, report_metadata: dict = None):
    """
    Save extracted financial statement data to an Excel file.

    Includes CoA columns (Code, Name, Category, Match Type, Confidence,
    Reasoning, Needs Review, Citation) if categorization metadata is present.

    Optionally appends a "Quality Report" sheet when `report_metadata` is provided.

    Args:
        data: Dict with keys: title, periods, sections (rows may have 'categorization')
        output_path: Path to save the Excel file
        include_coa_columns: Whether to add CoA columns for categorized data
        report_metadata: Optional dict with evaluation results, guardrail flags, etc.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = data.get("statement_type", "Statement")

    # Styles
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E79")
    coa_header_font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    coa_header_fill = PatternFill("solid", start_color="5B9BD5")
    section_font = Font(name="Arial", bold=True, size=10, color="1F4E79")
    section_fill = PatternFill("solid", start_color="D6E4F0")
    subtotal_font = Font(name="Arial", bold=True, size=10)
    subtotal_fill = PatternFill("solid", start_color="EBF5FB")
    normal_font = Font(name="Arial", size=10)
    coa_font = Font(name="Arial", size=9, color="006600")
    coa_fill = PatternFill("solid", start_color="E6F7FF")
    review_font = Font(name="Arial", size=9, color="CC0000")
    review_fill = PatternFill("solid", start_color="FFF0F0")
    warn_font = Font(name="Arial", size=9, color="CC6600")
    warn_fill = PatternFill("solid", start_color="FFF8E1")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        bottom=Side(border_style="thin", color="AAAAAA")
    )

    periods = data.get("periods", [])
    num_periods = len(periods)

    # Check if any rows have categorization
    has_categorization = any(
        "categorization" in row
        for section in data.get("sections", [])
        for row in section.get("rows", [])
        if not row.get("is_subtotal")
    )

    # CoA columns (placed to the RIGHT of period columns)
    coa_cols = ["CoA Code", "CoA Name", "CoA Category", "Match Type",
                "Confidence", "Reasoning", "Needs Review", "Citation"]
    show_coa = include_coa_columns and has_categorization
    num_coa_cols = len(coa_cols) if show_coa else 0

    # Layout: Line Item | Period 1 | Period 2 | ... | CoA Code | CoA Name | ...
    # Period columns start at col 2, CoA columns start after periods
    period_start_col = 2
    coa_start_col = period_start_col + num_periods
    total_cols = 1 + num_periods + num_coa_cols

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=data.get("title", "Financial Statement"))
    title_cell.font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 22

    # Row 2: column headers
    ws.cell(row=2, column=1, value="Line Item").font = header_font
    ws.cell(row=2, column=1).fill = header_fill

    # Period headers (columns 2 onward)
    for i, period in enumerate(periods, start=period_start_col):
        cell = ws.cell(row=2, column=i, value=period)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # CoA column headers (after period columns)
    for offset, header in enumerate(coa_cols):
        if show_coa:
            col = coa_start_col + offset
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = coa_header_font
            cell.fill = coa_header_fill
            cell.alignment = center_align
    ws.row_dimensions[2].height = 18

    current_row = 3

    for section in data.get("sections", []):
        # Section header
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=total_cols)
        sec_cell = ws.cell(row=current_row, column=1, value=section["name"])
        sec_cell.font = section_font
        sec_cell.fill = section_fill
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        for row in section.get("rows", []):
            # Label column
            label_cell = ws.cell(row=current_row, column=1, value=row["label"])
            label_cell.font = subtotal_font if row.get("is_subtotal") else normal_font
            if row.get("is_subtotal"):
                label_cell.fill = subtotal_fill
                label_cell.border = thin_border

            # Value columns (period values)
            for i, val in enumerate(row.get("values", []), start=period_start_col):
                val_cell = ws.cell(row=current_row, column=i, value=val)
                val_cell.font = subtotal_font if row.get("is_subtotal") else normal_font
                val_cell.alignment = right_align
                if row.get("is_subtotal"):
                    val_cell.fill = subtotal_fill
                    val_cell.border = thin_border

            # CoA columns (to the right of period values)
            if show_coa and not row.get("is_subtotal"):
                cat = row.get("categorization", {})
                is_unmatched = cat.get("match_type") == "unmatched"
                is_needs_review = cat.get("needs_review", False)
                is_low_conf = cat.get("confidence") in ("low", "unmatched")
                is_section_hdr = cat.get("match_type") == "section_header"

                # Determine row style based on status
                if is_unmatched or is_needs_review:
                    row_font = review_font
                    row_fill = review_fill
                elif is_low_conf and not is_section_hdr:
                    row_font = warn_font
                    row_fill = warn_fill
                else:
                    row_font = coa_font
                    row_fill = coa_fill

                coa_values = [
                    cat.get("coa_code", ""),
                    cat.get("coa_name", ""),
                    cat.get("coa_category", ""),
                    cat.get("match_type", ""),
                    cat.get("confidence", ""),
                    cat.get("reasoning", ""),
                    "YES" if is_needs_review else "No",
                    cat.get("citation", ""),
                ]
                for offset, val in enumerate(coa_values):
                    col = coa_start_col + offset
                    cell = ws.cell(row=current_row, column=col, value=val)
                    cell.font = row_font
                    cell.fill = row_fill
                    cell.alignment = left_align if offset >= 5 else center_align

            current_row += 1

        current_row += 1  # blank row between sections

    # Column widths
    ws.column_dimensions["A"].width = 40
    for i in range(period_start_col, period_start_col + num_periods):
        col_letter = utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = 18
    if show_coa:
        for offset in range(num_coa_cols):
            col_letter = utils.get_column_letter(coa_start_col + offset)
            widths = [12, 35, 18, 14, 12, 40, 14, 28]
            ws.column_dimensions[col_letter].width = widths[offset]

    # Phase 4 — append Quality Report sheet if metadata provided
    if report_metadata:
        _write_quality_report_sheet(wb, report_metadata)

    wb.save(output_path)
